#!/usr/bin/env python3
"""One-off importer: pulls Standard Market 2026-27 rates out of the SharePoint
'Master list 2026-27.xlsx' pricing workbook and writes prices/<id>-2027.json
for every route that already has a live product page on the site.

Source workbook: FIT/Packages-Products/2026-27/Master list 2026-27.xlsx (SharePoint).
Only the "Standard Market" columns are used (see chat decision on price tier).
A style's day-by-day Excel section that is blank (no Day #/component text) is
treated as a copy-paste relic with no real product behind it and is skipped,
even if the Menu sheet's checkbox or a computed rate table says otherwise.

This script does not touch products/*.json — every existing route's set of
travel styles already matches the workbook's populated (non-relic) sections
exactly (verified 0 mismatches across all 23 live routes before writing this).
"""
import json
import os
import sys
from datetime import datetime

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = sys.argv[1] if len(sys.argv) > 1 else None

STYLE_MAP = {"Regular FIT": "trains", "Private tour": "private", "Self Drive": "selfdrive"}

# The 8 Finland/Norway/Sweden winter & Xmas route sheets whose Regular FIT
# Start/End date *cells* are a full year ahead of every other sheet (a stale
# copy-paste - confirmed against each sheet's own Self Drive block and every
# other route, which all agree on Nov 2026 - Nov 2027). Only 5 of the 8 have
# a live product page today; the date-shift correction below applies to
# whichever of these are actually parsed.
NORDIC_CORRUPTED_DATE_ROUTES = {"10.1", "10.2", "10.3", "10.4", "10.5", "10.6", "11.1", "11.2"}

# These Arctic-winter routes (Tromso, Kiruna, Rovaniemi) only ever run in
# winter - the "summer" row present in their Regular FIT season table is a
# leftover template row, not a real second product, and must be dropped
# rather than shown as a second season (confirmed by the user, who sells
# these routes and knows they never run in summer).
WINTER_ONLY_ROUTES = {"10.1", "10.2", "10.3", "10.4", "10.5", "10.6"}

# The private-tour Min-Pax tables' own "summer window" date label is a typo
# repeated verbatim across all 23 sheets ("01.04.2026 - 30.11.2027" - a
# 20-month span, one year earlier at the start than every Regular FIT/Self
# Drive season table's own Apr 2027 start for the same 2026-27 rate cycle).
# Rather than reproduce that typo, private-tour validity uses the same
# Nov 2026 - Mar 2027 / Apr 2027 - Nov 2027 cycle every other table in the
# workbook agrees on.
PRIVATE_TOUR_WINDOWS = {
    "winter": (datetime(2026, 11, 1), datetime(2027, 3, 31)),
    "summer": (datetime(2027, 4, 1), datetime(2027, 11, 30)),
}


def fmt_date(dt):
    return dt.strftime("%-d %b %Y")


def shift_years(dt, years):
    if years == 0:
        return dt
    try:
        return dt.replace(year=dt.year + years)
    except ValueError:
        return dt.replace(month=2, day=28, year=dt.year + years)

# Excel sheet name -> repo product id (only "1.5" differs: repo file is ireland-discovery)
SHEET_TO_PRODUCT_ID = {
    "1.1": "1.1", "1.2": "1.2", "1.3": "1.3", "1.4": "1.4", "1.5": "ireland-discovery",
    "2.2": "2.2", "2.3": "2.3", "2.4": "2.4", "2.5": "2.5", "2.6": "2.6",
    "3.1": "3.1", "3.2": "3.2", "4.1": "4.1", "4.2": "4.2",
    "5.1": "5.1", "5.2": "5.2", "6.1": "6.1", "9.1": "9.1",
    "10.1": "10.1", "10.3": "10.3", "10.4": "10.4", "10.5": "10.5", "10.6": "10.6",
}


def cellstr(ws, r, c):
    v = ws.cell(r, c).value
    return v.strip() if isinstance(v, str) else v


def money(v):
    return None if v is None else round(float(v))


def find_component_sections(ws):
    """Return [(row, style_label, row_range_hint)] plus the Optional-tours row."""
    comp_rows = []
    for r in range(1, ws.max_row + 1):
        v = cellstr(ws, r, 2)
        if isinstance(v, str) and v.startswith("Component - ") and v != "Component - Hotels":
            comp_rows.append((r, v.replace("Component - ", "").strip()))
    opt_row = None
    for r in range(1, ws.max_row + 1):
        if cellstr(ws, r, 2) == "Optional" and cellstr(ws, r, 3) == "Price AD":
            opt_row = r
            break
    return comp_rows, opt_row


def section_has_itinerary(ws, r0, r1):
    for r in range(r0 + 1, r1):
        a, b = cellstr(ws, r, 1), cellstr(ws, r, 2)
        if (isinstance(a, str) and a.strip()) or (isinstance(b, str) and b.strip() and b.strip() != "Optional"):
            return True
    return False


def parse_season_table(ws, header_row, start_col_hint):
    for c in range(max(1, start_col_hint - 3), start_col_hint + 15):
        if cellstr(ws, header_row, c) == "Start" and cellstr(ws, header_row, c + 1) == "End":
            cols = {"start": c, "end": c + 1, "s3": c + 2, "t3": c + 3, "c3": c + 4, "s4": c + 5, "t4": c + 6, "c4": c + 7}
            rows = []
            for r in (header_row + 1, header_row + 2):
                d_start = ws.cell(r, cols["start"]).value
                d_end = ws.cell(r, cols["end"]).value
                if not isinstance(d_start, datetime):
                    continue
                rows.append({
                    "start": d_start, "end": d_end,
                    "3": {"single": ws.cell(r, cols["s3"]).value, "twin": ws.cell(r, cols["t3"]).value, "child": ws.cell(r, cols["c3"]).value},
                    "4": {"single": ws.cell(r, cols["s4"]).value, "twin": ws.cell(r, cols["t4"]).value, "child": ws.cell(r, cols["c4"]).value},
                })
            return rows
    return None


def parse_paxtier_table(ws, header_row, minpax_col):
    cols = {"pax": minpax_col, "s3": minpax_col + 1, "s4": minpax_col + 2, "w3": minpax_col + 3, "w4": minpax_col + 4}
    rows, r = [], header_row + 1
    while isinstance(ws.cell(r, cols["pax"]).value, (int, float)):
        rows.append({
            "pax": int(ws.cell(r, cols["pax"]).value),
            "window1": {"3star": ws.cell(r, cols["s3"]).value, "4star": ws.cell(r, cols["s4"]).value},
            "window2": {"3star": ws.cell(r, cols["w3"]).value, "4star": ws.cell(r, cols["w4"]).value},
        })
        r += 1
    return rows


def find_all_season_standard_tables(ws):
    """All STANDARD-market season (Start/End) tables in the sheet, in row order."""
    out = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if cellstr(ws, r, c) == "Start" and cellstr(ws, r, c + 1) == "End":
                if nearest_tier_marker(ws, r, c) == "STANDARD MARKET":
                    rows = parse_season_table(ws, r, c)
                    if rows:
                        out.append((r, rows))
    return out


def nearest_tier_marker(ws, r, c):
    """Classify column c as belonging to whichever of PREMIUM/STANDARD MARKET
    sits at the largest marker-column <= c, searching nearby rows above r.
    Two market groups sit side by side in the same row (e.g. col 7 = Premium,
    col 13 = Standard) so this must be nearest-to-the-left, not "any nearby
    occurrence" - a wide any-occurrence window bleeds into the other group."""
    for rr in range(r - 1, max(0, r - 4), -1):
        markers = []
        for cc in range(1, ws.max_column + 1):
            v = cellstr(ws, rr, cc)
            if v in ("PREMIUM MARKET", "STANDARD MARKET"):
                markers.append((cc, v))
        if markers:
            candidates = [(cc, v) for cc, v in markers if cc <= c]
            if candidates:
                return max(candidates, key=lambda x: x[0])[1]
    return None


def find_all_paxtier_standard_tables(ws):
    out = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if cellstr(ws, r, c) == "Min Pax":
                if nearest_tier_marker(ws, r, c) == "STANDARD MARKET":
                    out.append((r, parse_paxtier_table(ws, r, c)))
    return out


def route_currency(ws):
    for r in range(1, ws.max_row + 1):
        if cellstr(ws, r, 2) == "Component - Hotels":
            for rr in range(r + 1, r + 20):
                v = cellstr(ws, rr, 6)
                if v in ("GBP", "EUR"):
                    return {"GBP": "£", "EUR": "€"}[v]
    return "€"


def parse_route(ws, sheet_name):
    comp_rows, opt_row = find_component_sections(ws)
    ext = comp_rows + [(opt_row or ws.max_row + 1, None)]

    live_styles = []  # (style_key, kind) in document order
    for i, (r0, label) in enumerate(comp_rows):
        r1 = ext[i + 1][0]
        if section_has_itinerary(ws, r0, r1):
            live_styles.append(STYLE_MAP[label])

    season_tables = [rows for (_r, rows) in find_all_season_standard_tables(ws)]
    paxtier_tables = [rows for (_r, rows) in find_all_paxtier_standard_tables(ws)]

    date_shift_years = -1 if sheet_name in NORDIC_CORRUPTED_DATE_ROUTES else 0
    drop_summer = sheet_name in WINTER_ONLY_ROUTES

    variants = {}
    all_windows = []  # (start, end) across every season kept, for the route-level bounds
    season_i = 0
    for style in live_styles:
        if style == "private":
            if not paxtier_tables:
                raise ValueError("private is live but no paxtier standard table found")
            rows = paxtier_tables[0]
            validity = {
                season: {"from": fmt_date(start), "to": fmt_date(end)}
                for season, (start, end) in PRIVATE_TOUR_WINDOWS.items()
            }
            all_windows.extend(PRIVATE_TOUR_WINDOWS.values())
            variants["private"] = {
                "paxTiers": {
                    "winter": [{"pax": t["pax"], "3star": money(t["window1"]["3star"]), "4star": money(t["window1"]["4star"])} for t in rows],
                    "summer": [{"pax": t["pax"], "3star": money(t["window2"]["3star"]), "4star": money(t["window2"]["4star"])} for t in rows],
                },
                "validity": validity,
            }
        else:
            if season_i >= len(season_tables):
                raise ValueError(f"{style} is live but no season standard table left to assign")
            rows = season_tables[season_i]
            season_i += 1
            variant = {"3": {}, "4": {}}
            validity = {}
            for row in rows:
                start = shift_years(row["start"], date_shift_years)
                end = shift_years(row["end"], date_shift_years)
                # winter window starts in Nov, summer window starts in Apr
                season = "winter" if start.month in (10, 11, 12) else "summer"
                if season == "summer" and drop_summer:
                    continue
                for cat in ("3", "4"):
                    variant[cat][season] = {
                        "single": money(row[cat]["single"]),
                        "twin": money(row[cat]["twin"]),
                        "child": money(row[cat]["child"]),
                    }
                validity[season] = {"from": fmt_date(start), "to": fmt_date(end)}
                all_windows.append((start, end))
            variant["validity"] = validity
            variants[style] = variant

    optionals = []
    if opt_row:
        r = opt_row + 1
        while True:
            name, price = ws.cell(r, 2).value, ws.cell(r, 3).value
            if name is None:
                break
            if isinstance(price, (int, float)):
                optionals.append({"name": name.strip() if isinstance(name, str) else name, "price": money(price)})
            r += 1

    # Route-level validFrom/validTo is only used where there's no season
    # selection to key off (e.g. the destination-index card blurb) - the
    # outer bound across every season window actually kept above.
    valid_from = fmt_date(min(w[0] for w in all_windows))
    valid_to = fmt_date(max(w[1] for w in all_windows))
    return {
        "validFrom": valid_from,
        "validTo": valid_to,
        "currency": route_currency(ws),
        "variants": variants,
        "optionalTours": optionals,
    }, live_styles


def main():
    if not XLSX:
        print("usage: import_prices_2026_27.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    report = []
    for sheet_name, product_id in SHEET_TO_PRODUCT_ID.items():
        ws = wb[sheet_name]
        data, live_styles = parse_route(ws, sheet_name)

        product_path = os.path.join(REPO, "products", f"{product_id}.json")
        product = json.load(open(product_path))
        repo_styles = sorted(product["styles"].keys())
        if sorted(live_styles) != repo_styles:
            raise ValueError(f"{sheet_name}: excel live styles {sorted(live_styles)} != repo styles {repo_styles}")

        out_path = os.path.join(REPO, "prices", f"{product_id}-2027.json")
        with open(out_path, "w") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
            f.write("\n")
        report.append((sheet_name, product_id, live_styles, out_path))

    print(f"Wrote {len(report)} price files:")
    for sheet_name, product_id, styles, path in report:
        print(f"  {sheet_name:6s} -> {os.path.relpath(path, REPO):30s} styles={styles}")


if __name__ == "__main__":
    main()
